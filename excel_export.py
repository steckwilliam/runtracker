"""Build styled RunTracker Excel workbooks for download."""

from datetime import datetime, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from database import get_dashboard_runs, moving_time_to_seconds, pace_to_seconds

HEADER_FILL = PatternFill("solid", fgColor="2563EB")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)

FILENAME_BY_RANGE = {
    "30d": "RunTracker_30.xlsx",
    "90d": "RunTracker_90.xlsx",
    "365d": "RunTracker_365.xlsx",
    "all": "RunTracker_All.xlsx",
}

TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)

HEADERS = ("Date", "Start Time", "Distance", "Pace", "Time", "Weather")
COLUMN_WIDTHS = (12, 12, 11, 10, 11, 22)


def build_excel_export(range_key=None):
    """Return (bytes, filename, resolved_range_key) for the selected date range."""
    runs, resolved = get_dashboard_runs(range_key)

    wb = Workbook()
    _build_runs_sheet(wb.active, runs)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = FILENAME_BY_RANGE.get(resolved, f"RunTracker_{resolved}.xlsx")
    return buffer.getvalue(), filename, resolved


def _build_runs_sheet(ws, runs):
    ws.title = "Runs"
    ws.append(list(HEADERS))
    _style_header_row(ws, 1, len(HEADERS))

    for run in runs:
        ws.append(_run_row_values(run))

    last_row = max(len(runs) + 1, 1)
    if not runs:
        ws.append([None] * len(HEADERS))
        last_row = 2

    for row_idx in range(2, last_row + 1):
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row_idx, col)
            cell.border = THIN_BORDER

        distance_cell = ws.cell(row_idx, 3)
        if isinstance(distance_cell.value, (int, float)):
            distance_cell.number_format = "0.0"

        pace_cell = ws.cell(row_idx, 4)
        if isinstance(pace_cell.value, timedelta):
            pace_cell.number_format = "m:ss"

        time_cell = ws.cell(row_idx, 5)
        if isinstance(time_cell.value, timedelta):
            time_cell.number_format = "[h]:mm:ss"

    table = Table(displayName="RunLog", ref=f"A1:F{last_row}")
    table.tableStyleInfo = TABLE_STYLE
    ws.add_table(table)

    for idx, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"


def _run_row_values(run):
    run_date = datetime.strptime(run["date"], "%Y-%m-%d").date()
    pace_seconds = run.get("pace_seconds")
    if pace_seconds is None:
        pace_seconds = pace_to_seconds(run["pace_per_mile"])

    moving_time = run.get("moving_time") or ""
    time_value = (
        timedelta(seconds=moving_time_to_seconds(moving_time)) if moving_time else None
    )

    return [
        run_date,
        run.get("start_time_display") or "—",
        round(float(run["distance_miles"]), 2),
        timedelta(seconds=int(pace_seconds)),
        time_value,
        _format_weather_for_export(run),
    ]


def _format_weather_for_export(run):
    """Weather text without emoji icons (temp + condition only)."""
    if run.get("temperature_f") is None and not run.get("weather_condition"):
        return "—"
    parts = []
    if run.get("temperature_f") is not None:
        parts.append(f"{round(run['temperature_f'])}°F")
    if run.get("weather_condition"):
        parts.append(run["weather_condition"])
    return " ".join(parts) if parts else "—"


def _style_header_row(ws, row_num, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row_num, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER
